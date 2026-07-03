import os
import io
import json
from functools import wraps
from flask import Flask, jsonify, request, send_file, send_from_directory, session, redirect
from db import get_conn, init_db, DB_PATH, QUARTERS, next_fy_label
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config

BASE = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, static_folder='static')
app.secret_key = os.environ.get('APPT_SECRET_KEY', config.SECRET_KEY)
LOGIN_PASSWORD = os.environ.get('APPT_PASSWORD', config.LOGIN_PASSWORD)
ADMIN_PASSWORD = os.environ.get('APPT_ADMIN_PASSWORD', getattr(config, 'ADMIN_PASSWORD', 'admin'))
app.config['PERMANENT_SESSION_LIFETIME'] = 60 * 60 * 24 * 30  # stay signed in 30 days
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # never let browsers cache static files stale after an update

init_db()

def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'not authenticated'}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return wrapped

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return send_from_directory(app.static_folder, 'login.html')
    data = request.get_json(force=True) if request.is_json else request.form
    if data.get('password') == LOGIN_PASSWORD:
        session['logged_in'] = True
        session.permanent = True
        return jsonify({'status': 'ok'})
    return jsonify({'error': 'Incorrect password'}), 401

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'status': 'ok'})

# ---------------------------------------------------------------- helpers --

def row_to_dict(row):
    return {k: row[k] for k in row.keys()}

def q_sum(r, prefix=''):
    return (r[prefix + 'q1'] or 0) + (r[prefix + 'q2'] or 0) + (r[prefix + 'q3'] or 0) + (r[prefix + 'q4'] or 0)

def get_workflow(conn):
    ws = row_to_dict(conn.execute("SELECT * FROM workflow_state WHERE id=1").fetchone())
    fys = [row_to_dict(r) for r in conn.execute("SELECT * FROM financial_years ORDER BY start_year")]
    frozen = {}
    for r in conn.execute("SELECT fy, quarter FROM quarter_freeze"):
        frozen.setdefault(r['fy'], []).append(r['quarter'])
    return {
        'current_fy': ws['current_fy'],
        'current_quarter': ws['current_quarter'],
        'financial_years': fys,
        'frozen': frozen,
    }

def field_permissions(workflow, fy):
    """For a given fy being viewed, return which plan/achv fields are editable per quarter."""
    perms = {}
    is_current_fy = (fy == workflow['current_fy'])
    cur_idx = QUARTERS.index(workflow['current_quarter']) + 1
    frozen_q = set(workflow['frozen'].get(fy, []))
    for i, q in enumerate(QUARTERS, start=1):
        frozen = q in frozen_q
        plan_ok = is_current_fy and not frozen
        achv_ok = is_current_fy and not frozen and i <= cur_idx
        perms[q] = {'plan': plan_ok, 'achv': achv_ok, 'frozen': frozen}
    return perms

def get_all_blocks(conn):
    """id -> {district, name, category}"""
    return {r['id']: row_to_dict(r) for r in conn.execute("SELECT id, district, name, category FROM blocks")}

def build_entry_index(conn, district=None):
    """key: (indicator_code, district, block_id, level) -> {fy: row}"""
    q = "SELECT * FROM entries"
    params = []
    if district:
        q += " WHERE district = ?"
        params.append(district)
    idx = {}
    for r in conn.execute(q, params):
        key = (r['indicator_code'], r['district'], r['block_id'], r['level'])
        idx.setdefault(key, {})[r['fy']] = row_to_dict(r)
    return idx

def carry_forward_achieved(entry_index, fy_order, is_cumulative, key, target_fy):
    """Total achieved for target_fy, including carried-forward totals from prior years
    (plus any manually-seeded baseline) for cumulative indicators."""
    by_fy = entry_index.get(key, {})
    total = 0
    if is_cumulative:
        for fy in fy_order:
            if fy == target_fy:
                break
            r = by_fy.get(fy)
            if r:
                total += q_sum(r) + (r['manual_opening'] or 0)
    target_row = by_fy.get(target_fy)
    this_year = q_sum(target_row) if target_row else 0
    if is_cumulative and target_row and fy_order and target_fy == fy_order[0]:
        total += target_row['manual_opening'] or 0
    return total, this_year, total + this_year

def opening_balance_editable(workflow, fy, is_cumulative):
    fy_order = [f['fy'] for f in workflow['financial_years']]
    is_base_fy = bool(fy_order) and fy == fy_order[0]
    return is_cumulative and is_base_fy and fy == workflow['current_fy']

def plan_for_fy(entry_index, key, target_fy):
    r = entry_index.get(key, {}).get(target_fy)
    if not r:
        return 0
    return (r['q1_plan'] or 0) + (r['q2_plan'] or 0) + (r['q3_plan'] or 0) + (r['q4_plan'] or 0)

# ------------------------------------------------------------------ pages --

@app.route('/')
@login_required
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/static/<path:path>')
def statics(path):
    return send_from_directory(app.static_folder, path)

# ------------------------------------------------------------------- meta --

@app.route('/api/meta')
@login_required
def meta():
    conn = get_conn()
    districts = [r['name'] for r in conn.execute("SELECT name FROM districts ORDER BY name")]
    blocks = {}
    for r in conn.execute("SELECT id, district, name, category FROM blocks ORDER BY district, category, name"):
        blocks.setdefault(r['district'], []).append({'id': r['id'], 'name': r['name'], 'category': r['category']})
    indicators = [row_to_dict(r) for r in conn.execute(
        "SELECT code, theme, subtheme, text, period FROM indicators ORDER BY theme, code")]
    themes = sorted(set(i['theme'] for i in indicators))
    workflow = get_workflow(conn)
    conn.close()
    return jsonify({'districts': districts, 'blocks': blocks, 'indicators': indicators, 'themes': themes,
                     'workflow': workflow})

# --------------------------------------------------------------- workflow --

@app.route('/api/workflow')
@login_required
def workflow_get():
    conn = get_conn()
    w = get_workflow(conn)
    conn.close()
    return jsonify(w)

@app.route('/api/workflow/freeze', methods=['POST'])
@login_required
def workflow_freeze():
    data = request.get_json(force=True)
    if data.get('admin_password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Incorrect admin password.'}), 403

    fy = data.get('fy')
    quarter = data.get('quarter')

    conn = get_conn()
    ws = row_to_dict(conn.execute("SELECT * FROM workflow_state WHERE id=1").fetchone())
    if fy != ws['current_fy'] or quarter != ws['current_quarter']:
        conn.close()
        return jsonify({'error': 'Only the current open quarter can be frozen.'}), 400

    already = conn.execute("SELECT 1 FROM quarter_freeze WHERE fy=? AND quarter=?", (fy, quarter)).fetchone()
    if already:
        conn.close()
        return jsonify({'error': 'This quarter is already frozen.'}), 400

    conn.execute("INSERT INTO quarter_freeze (fy, quarter) VALUES (?,?)", (fy, quarter))

    idx = QUARTERS.index(quarter)
    if idx < len(QUARTERS) - 1:
        new_fy, new_quarter = fy, QUARTERS[idx + 1]
    else:
        new_fy = next_fy_label(fy)
        new_quarter = QUARTERS[0]
        conn.execute("INSERT OR IGNORE INTO financial_years (fy, start_year) VALUES (?,?)",
                     (new_fy, int(new_fy.split('-')[0])))

    conn.execute("UPDATE workflow_state SET current_fy=?, current_quarter=? WHERE id=1", (new_fy, new_quarter))
    conn.commit()
    w = get_workflow(conn)
    conn.close()
    return jsonify({'status': 'ok', 'workflow': w})

@app.route('/api/workflow/unfreeze', methods=['POST'])
@login_required
def workflow_unfreeze():
    """Reopens the most recently frozen quarter (reverses exactly one freeze step)."""
    data = request.get_json(force=True)
    if data.get('admin_password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Incorrect admin password.'}), 403

    conn = get_conn()
    ws = row_to_dict(conn.execute("SELECT * FROM workflow_state WHERE id=1").fetchone())
    cur_fy, cur_q = ws['current_fy'], ws['current_quarter']

    idx = QUARTERS.index(cur_q)
    if idx > 0:
        prev_fy, prev_q = cur_fy, QUARTERS[idx - 1]
    else:
        fys = [f['fy'] for f in conn.execute("SELECT fy FROM financial_years ORDER BY start_year")]
        cur_pos = fys.index(cur_fy) if cur_fy in fys else -1
        if cur_pos <= 0:
            conn.close()
            return jsonify({'error': 'Nothing to unfreeze — this is the very first period.'}), 400
        prev_fy, prev_q = fys[cur_pos - 1], QUARTERS[-1]

    frozen = conn.execute("SELECT 1 FROM quarter_freeze WHERE fy=? AND quarter=?", (prev_fy, prev_q)).fetchone()
    if not frozen:
        conn.close()
        return jsonify({'error': 'The previous period is not frozen — nothing to unfreeze.'}), 400

    conn.execute("DELETE FROM quarter_freeze WHERE fy=? AND quarter=?", (prev_fy, prev_q))
    conn.execute("UPDATE workflow_state SET current_fy=?, current_quarter=? WHERE id=1", (prev_fy, prev_q))
    conn.commit()
    w = get_workflow(conn)
    conn.close()
    return jsonify({'status': 'ok', 'workflow': w})

# ---------------------------------------------------------------- entries --

@app.route('/api/entries')
@login_required
def get_entries():
    """Fetch entries (with carry-forward + edit permissions) for a district/block/fy.
    The block's Grassroot/Ecosystem category is looked up server-side — the client just picks a block."""
    district = request.args.get('district')
    block_id = request.args.get('block_id')
    theme = request.args.get('theme')

    if not district or not block_id:
        return jsonify({'error': 'district and block_id are required'}), 400
    block_id_val = int(block_id)

    conn = get_conn()
    workflow = get_workflow(conn)
    fy = request.args.get('fy') or workflow['current_fy']
    fy_order = [f['fy'] for f in workflow['financial_years']]

    block = conn.execute("SELECT id, district, name, category FROM blocks WHERE id=?", (block_id_val,)).fetchone()
    if not block:
        conn.close()
        return jsonify({'error': 'Unknown block'}), 404
    level = block['category'].lower()

    indicators = [row_to_dict(r) for r in conn.execute(
        "SELECT code, theme, subtheme, text, period FROM indicators ORDER BY theme, code")]
    if theme:
        indicators = [i for i in indicators if i['theme'] == theme]

    entry_index = build_entry_index(conn, district=district)
    conn.close()

    perms = field_permissions(workflow, fy)

    rows = []
    for ind in indicators:
        key = (ind['code'], district, block_id_val, level)
        by_fy = entry_index.get(key, {})
        r = by_fy.get(fy, {})
        is_cumulative = (ind['period'] == 'Cumulative')
        opening, this_year, total = carry_forward_achieved(entry_index, fy_order, is_cumulative, key, fy)
        rows.append({
            'indicator_code': ind['code'], 'theme': ind['theme'], 'subtheme': ind['subtheme'],
            'text': ind['text'], 'period': ind['period'],
            'fy': fy,
            'q1_plan': r.get('q1_plan', 0), 'q2_plan': r.get('q2_plan', 0),
            'q3_plan': r.get('q3_plan', 0), 'q4_plan': r.get('q4_plan', 0),
            'q1': r.get('q1', 0), 'q2': r.get('q2', 0), 'q3': r.get('q3', 0), 'q4': r.get('q4', 0),
            'manual_opening': r.get('manual_opening', 0),
            'opening_editable': opening_balance_editable(workflow, fy, is_cumulative),
            'opening_balance': opening,
            'this_year_achieved': this_year,
            'total_achieved': total,
        })

    return jsonify({'fy': fy, 'block_category': block['category'], 'field_permissions': perms, 'entries': rows})

@app.route('/api/entries/bulk', methods=['POST'])
@login_required
def save_entries_bulk():
    """
    Body: { district, block_id, fy, entries: [{indicator_code, q1_plan..q4_plan, q1..q4}, ...] }
    The block's category (Grassroot/Ecosystem) is looked up server-side and used as 'level' —
    the client never chooses this directly. Server enforces freeze / editability rules —
    disallowed fields are silently kept at their old value.
    """
    data = request.get_json(force=True)
    district = data['district']
    block_id = data.get('block_id')
    entries = data.get('entries', [])

    if not block_id:
        return jsonify({'error': 'block_id is required'}), 400

    conn = get_conn()
    block = conn.execute("SELECT category FROM blocks WHERE id=? AND district=?", (block_id, district)).fetchone()
    if not block:
        conn.close()
        return jsonify({'error': 'Unknown block for this district'}), 404
    level = block['category'].lower()

    workflow = get_workflow(conn)
    fy = data.get('fy') or workflow['current_fy']
    perms = field_permissions(workflow, fy)
    ind_periods = {r['code']: r['period'] for r in conn.execute("SELECT code, period FROM indicators")}

    saved = 0
    for e in entries:
        code = e['indicator_code']
        existing = conn.execute("""
            SELECT * FROM entries WHERE indicator_code=? AND district=? AND block_id=? AND fy=?
        """, (code, district, block_id, fy)).fetchone()
        existing = row_to_dict(existing) if existing else {}

        vals = {}
        for i, q in enumerate(QUARTERS, start=1):
            qn = q.lower()
            plan_field = f'{qn}_plan'
            achv_field = qn
            if perms[q]['plan']:
                vals[plan_field] = e.get(plan_field, existing.get(plan_field, 0)) or 0
            else:
                vals[plan_field] = existing.get(plan_field, 0) or 0
            if perms[q]['achv']:
                vals[achv_field] = e.get(achv_field, existing.get(achv_field, 0)) or 0
            else:
                vals[achv_field] = existing.get(achv_field, 0) or 0

        is_cum = ind_periods.get(code) == 'Cumulative'
        if opening_balance_editable(workflow, fy, is_cum):
            vals['manual_opening'] = e.get('manual_opening', existing.get('manual_opening', 0)) or 0
        else:
            vals['manual_opening'] = existing.get('manual_opening', 0) or 0

        conn.execute("""
            INSERT INTO entries (indicator_code, district, block_id, level, fy,
                                  manual_opening, q1_plan, q2_plan, q3_plan, q4_plan, q1, q2, q3, q4, updated_at)
            VALUES (?,?,?,?,?, ?, ?,?,?,?, ?,?,?,?, CURRENT_TIMESTAMP)
            ON CONFLICT(indicator_code, district, block_id, fy) DO UPDATE SET
                level=excluded.level,
                manual_opening=excluded.manual_opening,
                q1_plan=excluded.q1_plan, q2_plan=excluded.q2_plan, q3_plan=excluded.q3_plan, q4_plan=excluded.q4_plan,
                q1=excluded.q1, q2=excluded.q2, q3=excluded.q3, q4=excluded.q4,
                updated_at=CURRENT_TIMESTAMP
        """, (
            code, district, block_id, level, fy,
            vals['manual_opening'],
            vals['q1_plan'], vals['q2_plan'], vals['q3_plan'], vals['q4_plan'],
            vals['q1'], vals['q2'], vals['q3'], vals['q4']
        ))
        saved += 1
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'saved': saved, 'fy': fy, 'level': level})

# ------------------------------------------------------------ consolidation --

@app.route('/api/consolidation/district/<district>')
@login_required
def consolidation_district(district):
    """Block-wise (Grassroot + Ecosystem) + district-total breakdown for every indicator, for a given FY."""
    conn = get_conn()
    workflow = get_workflow(conn)
    fy = request.args.get('fy') or workflow['current_fy']
    fy_order = [f['fy'] for f in workflow['financial_years']]

    indicators = [row_to_dict(r) for r in conn.execute(
        "SELECT code, theme, subtheme, text, period FROM indicators ORDER BY theme, code")]
    blocks = [row_to_dict(r) for r in conn.execute(
        "SELECT id, name, category FROM blocks WHERE district=? ORDER BY category, name", (district,))]
    entry_index = build_entry_index(conn, district=district)
    conn.close()

    result = []
    for ind in indicators:
        code = ind['code']
        is_cum = ind['period'] == 'Cumulative'
        block_vals = []
        grassroot_plan = 0
        grassroot_achv = 0
        ecosystem_plan = 0
        ecosystem_achv = 0
        for b in blocks:
            level = b['category'].lower()
            key = (code, district, b['id'], level)
            plan = plan_for_fy(entry_index, key, fy)
            _, _, achv = carry_forward_achieved(entry_index, fy_order, is_cum, key, fy)
            block_vals.append({'block': b['name'], 'block_id': b['id'], 'category': b['category'],
                                'annual_plan': plan, 'achieved': achv})
            if b['category'] == 'Grassroot':
                grassroot_plan += plan
                grassroot_achv += achv
            else:
                ecosystem_plan += plan
                ecosystem_achv += achv
        result.append({
            **ind,
            'blocks': block_vals,
            'grassroot_plan': grassroot_plan,
            'grassroot_achieved': grassroot_achv,
            'ecosystem_plan': ecosystem_plan,
            'ecosystem_achieved': ecosystem_achv,
            'district_plan': grassroot_plan + ecosystem_plan,
            'district_achieved': grassroot_achv + ecosystem_achv,
        })
    return jsonify({'district': district, 'fy': fy, 'blocks': blocks, 'indicators': result})

@app.route('/api/consolidation/state')
@login_required
def consolidation_state():
    """District-wise + state-total roll up for every indicator, for a given FY, split Grassroot/Ecosystem."""
    conn = get_conn()
    workflow = get_workflow(conn)
    fy = request.args.get('fy') or workflow['current_fy']
    fy_order = [f['fy'] for f in workflow['financial_years']]

    indicators = [row_to_dict(r) for r in conn.execute(
        "SELECT code, theme, subtheme, text, period FROM indicators ORDER BY theme, code")]
    districts = [r['name'] for r in conn.execute("SELECT name FROM districts ORDER BY name")]
    blocks_by_district = {}
    for r in conn.execute("SELECT id, district, category FROM blocks"):
        blocks_by_district.setdefault(r['district'], []).append({'id': r['id'], 'category': r['category']})

    entry_index = build_entry_index(conn)
    conn.close()

    result = []
    for ind in indicators:
        code = ind['code']
        is_cum = ind['period'] == 'Cumulative'
        by_district = []
        state_plan = 0
        state_achv = 0
        state_grassroot_achv = 0
        state_ecosystem_achv = 0
        for dist in districts:
            dist_plan = 0
            dist_achv = 0
            dist_grassroot = 0
            dist_ecosystem = 0
            for b in blocks_by_district.get(dist, []):
                level = b['category'].lower()
                key = (code, dist, b['id'], level)
                p = plan_for_fy(entry_index, key, fy)
                _, _, a = carry_forward_achieved(entry_index, fy_order, is_cum, key, fy)
                dist_plan += p
                dist_achv += a
                if b['category'] == 'Grassroot':
                    dist_grassroot += a
                else:
                    dist_ecosystem += a
            by_district.append({'district': dist, 'annual_plan': dist_plan, 'achieved': dist_achv,
                                 'grassroot_achieved': dist_grassroot, 'ecosystem_achieved': dist_ecosystem})
            state_plan += dist_plan
            state_achv += dist_achv
            state_grassroot_achv += dist_grassroot
            state_ecosystem_achv += dist_ecosystem
        result.append({
            **ind,
            'by_district': by_district,
            'state_plan': state_plan,
            'state_achieved': state_achv,
            'state_grassroot_achieved': state_grassroot_achv,
            'state_ecosystem_achieved': state_ecosystem_achv,
        })
    return jsonify({'districts': districts, 'fy': fy, 'indicators': result})

# ----------------------------------------------------------------- dashboard --

@app.route('/api/dashboard')
@login_required
def dashboard():
    conn = get_conn()
    workflow = get_workflow(conn)
    fy = request.args.get('fy') or workflow['current_fy']
    fy_order = [f['fy'] for f in workflow['financial_years']]

    indicators = {r['code']: row_to_dict(r) for r in conn.execute("SELECT * FROM indicators")}
    blocks_meta = get_all_blocks(conn)
    entry_index = build_entry_index(conn)
    entries_this_fy_count = conn.execute("SELECT COUNT(*) c FROM entries WHERE fy=?", (fy,)).fetchone()['c']
    conn.close()

    theme_totals = {}
    district_totals = {}
    quarter_totals = {'q1': 0, 'q2': 0, 'q3': 0, 'q4': 0}
    block_heat = {}
    category_totals = {'Grassroot': {'plan': 0, 'achv': 0}, 'Ecosystem': {'plan': 0, 'achv': 0}}
    total_plan = 0
    total_achv = 0

    for key, by_fy in entry_index.items():
        code, district, block_id, level = key
        ind = indicators.get(code)
        if not ind:
            continue
        is_cum = ind['period'] == 'Cumulative'
        theme = ind['theme']
        plan = plan_for_fy(entry_index, key, fy)
        _, this_year, achv = carry_forward_achieved(entry_index, fy_order, is_cum, key, fy)
        r = by_fy.get(fy)

        total_plan += plan
        total_achv += achv

        t = theme_totals.setdefault(theme, {'plan': 0, 'achv': 0})
        t['plan'] += plan; t['achv'] += achv

        d = district_totals.setdefault(district, {'plan': 0, 'achv': 0})
        d['plan'] += plan; d['achv'] += achv

        block_meta = blocks_meta.get(block_id)
        category = block_meta['category'] if block_meta else ('Grassroot' if level == 'grassroot' else 'Ecosystem')
        ct = category_totals.setdefault(category, {'plan': 0, 'achv': 0})
        ct['plan'] += plan; ct['achv'] += achv

        if r:
            quarter_totals['q1'] += r['q1'] or 0
            quarter_totals['q2'] += r['q2'] or 0
            quarter_totals['q3'] += r['q3'] or 0
            quarter_totals['q4'] += r['q4'] or 0

        if block_id:
            bd = block_heat.setdefault(district, {})
            bb = bd.setdefault(str(block_id), {'plan': 0, 'achv': 0, 'category': category})
            bb['plan'] += plan; bb['achv'] += achv

    def pct(a, p):
        return round((a / p) * 100, 1) if p else 0

    theme_list = [{'theme': k, 'plan': v['plan'], 'achieved': v['achv'], 'pct': pct(v['achv'], v['plan'])}
                   for k, v in sorted(theme_totals.items())]
    district_list = [{'district': k, 'plan': v['plan'], 'achieved': v['achv'], 'pct': pct(v['achv'], v['plan'])}
                      for k, v in district_totals.items()]
    district_list.sort(key=lambda x: -x['pct'])

    return jsonify({
        'fy': fy,
        'overall': {'plan': total_plan, 'achieved': total_achv, 'pct': pct(total_achv, total_plan),
                    'entries_count': entries_this_fy_count},
        'by_theme': theme_list,
        'by_district': district_list,
        'quarterly': quarter_totals,
        'block_heat': block_heat,
        'grassroot_vs_ecosystem': {
            'grassroot': {'plan': category_totals['Grassroot']['plan'], 'achieved': category_totals['Grassroot']['achv'],
                          'pct': pct(category_totals['Grassroot']['achv'], category_totals['Grassroot']['plan'])},
            'ecosystem': {'plan': category_totals['Ecosystem']['plan'], 'achieved': category_totals['Ecosystem']['achv'],
                          'pct': pct(category_totals['Ecosystem']['achv'], category_totals['Ecosystem']['plan'])},
        },
    })

# ------------------------------------------------------------------- export --

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GRASSROOT_FILL = PatternFill(start_color="E8F6EF", end_color="E8F6EF", fill_type="solid")
ECOSYSTEM_FILL = PatternFill(start_color="F7FAFF", end_color="F7FAFF", fill_type="solid")
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

@app.route('/api/export/state')
@login_required
def export_state():
    fy = request.args.get('fy')
    with app.test_request_context(f'/api/consolidation/state?fy={fy}' if fy else '/api/consolidation/state'):
        session['logged_in'] = True
        resp = consolidation_state().get_json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "State Consolidation"

    headers = ['Code', 'Theme', 'Sub-theme', 'Indicator', 'Cumulative/Period'] + resp['districts'] + \
              ['STATE GRASSROOT', 'STATE ECOSYSTEM', 'STATE PLAN', 'STATE ACHIEVED', '% ACHIEVED']
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for ind in resp['indicators']:
        row = [ind['code'], ind['theme'], ind.get('subtheme') or '', ind['text'], ind.get('period') or '']
        for d in ind['by_district']:
            row.append(d['achieved'])
        pct = round((ind['state_achieved'] / ind['state_plan']) * 100, 1) if ind['state_plan'] else 0
        row += [ind['state_grassroot_achieved'], ind['state_ecosystem_achieved'], ind['state_plan'], ind['state_achieved'], pct]
        ws.append(row)

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)
    ws.freeze_panes = 'F2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"APPT_State_Consolidation_{resp['fy']}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export/district/<district>')
@login_required
def export_district(district):
    fy = request.args.get('fy')
    path = f'/api/consolidation/district/{district}' + (f'?fy={fy}' if fy else '')
    with app.test_request_context(path):
        session['logged_in'] = True
        resp = consolidation_district(district).get_json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = district[:31]

    block_names = [f"{b['name']} ({b['category'][0]})" for b in resp['blocks']]
    headers = ['Code', 'Theme', 'Sub-theme', 'Indicator', 'Cumulative/Period'] + block_names + \
              ['GRASSROOT TOTAL', 'ECOSYSTEM TOTAL', 'DISTRICT TOTAL', 'DISTRICT PLAN', '% ACHIEVED']
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for ind in resp['indicators']:
        row = [ind['code'], ind['theme'], ind.get('subtheme') or '', ind['text'], ind.get('period') or '']
        for b in ind['blocks']:
            row.append(b['achieved'])
        pct = round((ind['district_achieved'] / ind['district_plan']) * 100, 1) if ind['district_plan'] else 0
        row += [ind['grassroot_achieved'], ind['ecosystem_achieved'], ind['district_achieved'], ind['district_plan'], pct]
        ws.append(row)

    # tint block columns by category so Grassroot vs Ecosystem is visible at a glance
    for i, b in enumerate(resp['blocks']):
        col = 6 + i  # after Code, Theme, Sub-theme, Indicator, Cumulative/Period
        fill = GRASSROOT_FILL if b['category'] == 'Grassroot' else ECOSYSTEM_FILL
        for r in range(2, len(resp['indicators']) + 2):
            ws.cell(row=r, column=col).fill = fill

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)
    ws.freeze_panes = 'F2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"APPT_{district}_Consolidation_{resp['fy']}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    print(f"DB at {DB_PATH}")
    try:
        app.run(host='127.0.0.1', port=5050, debug=False, threaded=True)
    except OSError as e:
        print("\n[ERROR] Could not start the server.")
        if 'Address already in use' in str(e) or 'WinError 10048' in str(e):
            print("Port 5050 is already being used by another program (maybe the app")
            print("is already running in another window? Check for another 'APPT Server' window).")
        else:
            print(str(e))
        input("\nPress Enter to close...")
